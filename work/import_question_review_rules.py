from __future__ import annotations

import ast
import json
import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "outputs" / "rag_assets"


def parse_image_list(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    try:
        parsed = ast.literal_eval(text)
        if isinstance(parsed, list):
            return [str(item) for item in parsed if str(item).strip()]
    except Exception:
        pass
    return re.findall(
        r"(?:Manual\d+|Camera|drill\d*|jetski|Security_Camera|air_conditioner|exercise_bikes|fitness_trackers|oven|fax|generator|Dish_washer|Blower)_[A-Za-z0-9]+",
        text,
    )


def row_to_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    data = {str(header): row[i] for i, header in enumerate(headers) if header is not None}
    return {
        "id": str(data.get("题号") or "").strip(),
        "question": str(data.get("问题") or "").strip(),
        "manual_id": str(data.get("手册ID") or "").strip(),
        "product": str(data.get("产品") or "").strip(),
        "language": str(data.get("语言") or "").strip(),
        "question_type": str(data.get("题型") or "").strip(),
        "risk": str(data.get("风险等级") or "").strip(),
        "action": str(data.get("建议动作") or "").strip(),
        "issue": str(data.get("问题说明") or "").strip(),
        "current_image_count": int(float(data.get("当前图片数") or 0)),
        "current_images": parse_image_list(data.get("当前图片数组")),
        "suggested_images": parse_image_list(data.get("建议图片数组")),
        "model_feedback": str(data.get("给大模型的反馈") or "").strip(),
        "answer_snippet": str(data.get("当前答案片段") or "").strip(),
        "caption_preview": str(data.get("图片caption预览") or "").strip(),
    }


def main() -> None:
    path = Path(os.environ.get("REVIEW_XLSX", ""))
    if not path.exists():
        raise SystemExit("Set REVIEW_XLSX to the review workbook path")
    wb = load_workbook(path, read_only=True, data_only=True)

    review_ws = wb.worksheets[1]
    headers = [str(review_ws.cell(1, c).value or "") for c in range(1, review_ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for raw_row in review_ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and str(value).strip() for value in raw_row):
            continue
        row = row_to_dict(headers, raw_row)
        if row["id"]:
            rows.append(row)

    patch_ws = wb.worksheets[3]
    patch_headers = [str(patch_ws.cell(1, c).value or "") for c in range(1, patch_ws.max_column + 1)]
    patches = []
    for raw_row in patch_ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and str(value).strip() for value in raw_row):
            continue
        data = {patch_headers[i]: raw_row[i] for i in range(len(patch_headers))}
        patches.append(
            {
                "scope": str(data.get("适用手册/范围") or "").strip(),
                "trigger": str(data.get("触发条件") or "").strip(),
                "instruction": str(data.get("规则补丁/给模型的指令") or "").strip(),
            }
        )

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with (OUT_DIR / "question_review_rules.jsonl").open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
    with (OUT_DIR / "question_review_global_patches.jsonl").open("w", encoding="utf-8") as f:
        for patch in patches:
            f.write(json.dumps(patch, ensure_ascii=False) + "\n")

    summary = {
        "review_rows": len(rows),
        "forced_image_rows": sum(1 for row in rows if row["suggested_images"]),
        "patch_rows": len(patches),
        "actions": {},
        "risks": {},
    }
    for row in rows:
        summary["actions"][row["action"]] = summary["actions"].get(row["action"], 0) + 1
        summary["risks"][row["risk"]] = summary["risks"].get(row["risk"], 0) + 1
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
