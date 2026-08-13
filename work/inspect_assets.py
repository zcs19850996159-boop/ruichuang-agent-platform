from __future__ import annotations

import ast
import json
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "work" / "data_unzip" / "data"
END_DIR = ROOT / "work" / "end_unzip" / "end"


def parse_manual(path: Path) -> dict:
    raw = path.read_text(encoding="utf-8")
    try:
        payload = ast.literal_eval(raw)
    except Exception as exc:
        return {
            "file": path.name,
            "parse_error": str(exc),
            "chars": len(raw),
            "pic_placeholders": raw.count("<PIC>"),
            "image_ids": None,
        }

    text, image_ids = payload
    return {
        "file": path.name,
        "chars": len(text),
        "lines": text.count("\n") + 1,
        "pic_placeholders": text.count("<PIC>"),
        "image_ids": len(image_ids),
        "first_ids": image_ids[:5],
        "last_ids": image_ids[-5:],
        "mismatch": text.count("<PIC>") != len(image_ids),
    }


def sheet_preview(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    sheets = []
    for ws in wb.worksheets:
        non_empty_rows = []
        max_r = 0
        max_c = 0
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            if any(cell is not None for cell in values):
                non_empty_rows.append(values)
                row_index = len(non_empty_rows)
                max_r = max(max_r, row_index)
                for idx, cell in enumerate(values, start=1):
                    if cell is not None:
                        max_c = max(max_c, idx)
        sheets.append(
            {
                "sheet": ws.title,
                "rows": len(non_empty_rows),
                "cols": max_c,
                "preview": non_empty_rows[:8],
            }
        )
    wb.close()
    return {"file": path.name, "sheets": sheets}


def compact_rows(rows: list[list], max_cols: int = 10) -> list[list]:
    out = []
    for row in rows:
        values = list(row[:max_cols])
        while values and values[-1] is None:
            values.pop()
        values = [cell[:120] if isinstance(cell, str) else cell for cell in values]
        out.append(values)
    return out


def main() -> None:
    manuals = [parse_manual(path) for path in sorted(DATA_DIR.glob("*.txt"))]
    xlsx = [sheet_preview(path) for path in sorted(END_DIR.glob("*.xlsx"))]

    print("MANUAL_SUMMARY")
    print(
        json.dumps(
            [
                {
                    "file": m["file"],
                    "chars": m.get("chars"),
                    "pic_placeholders": m.get("pic_placeholders"),
                    "image_ids": m.get("image_ids"),
                    "mismatch": m.get("mismatch", "parse_error"),
                    "parse_error": m.get("parse_error"),
                    "first_ids": m.get("first_ids"),
                }
                for m in manuals
            ],
            ensure_ascii=False,
            indent=2,
        )
    )

    print("\nXLSX_SUMMARY")
    summary = []
    for book in xlsx:
        summary.append(
            {
                "file": book["file"],
                "sheets": [
                    {
                        "sheet": s["sheet"],
                        "rows": s["rows"],
                        "cols": s["cols"],
                        "preview": compact_rows(s["preview"]),
                    }
                    for s in book["sheets"]
                ],
            }
        )
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    print("\nRISK_WORD_COUNTS")
    risk_terms = re.compile(r"风险|缺|错|多|少|补|无|重复|不一致|异常")
    risk_hits = []
    for path in sorted(END_DIR.glob("*.xlsx")):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        hit_count = 0
        examples = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str) and risk_terms.search(cell):
                        hit_count += 1
                        if len(examples) < 8:
                            examples.append(cell[:120])
        wb.close()
        risk_hits.append({"file": path.name, "risk_text_cells": hit_count, "examples": examples})
    print(json.dumps(risk_hits, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
