from __future__ import annotations

import ast
import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
ASSET_DIR = ROOT / "outputs" / "rag_assets"
DOWNLOADS = Path.home() / "Downloads"
REPORT_DIR = ROOT / "work" / "human_checked_examples_unzip"
IMAGE_ID_RE = re.compile(r"^[A-Za-z][A-Za-z0-9]*(?:_[A-Za-z0-9]+)+$")


def compact(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def unescape_unicode_literals(text: str) -> str:
    return re.sub(
        r"\\u([0-9a-fA-F]{4})",
        lambda match: chr(int(match.group(1), 16)),
        text or "",
    )


def read_sheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []
    header = [compact(value) for value in rows[0]]
    out: list[dict[str, Any]] = []
    for raw in rows[1:]:
        if not any(value is not None and compact(value) for value in raw):
            continue
        row = {header[i]: raw[i] if i < len(raw) else None for i in range(len(header)) if header[i]}
        out.append(row)
    return out


def parse_json_list(value: Any) -> list[str]:
    text = compact(value)
    if not text:
        return []
    match = re.search(r"\[[^\]]*\]", text)
    if not match:
        return []
    try:
        parsed = json.loads(match.group(0))
    except Exception:
        try:
            parsed = ast.literal_eval(match.group(0))
        except Exception:
            return []
    if not isinstance(parsed, list):
        return []
    return [str(item).strip() for item in parsed if IMAGE_ID_RE.match(str(item).strip())]


def parse_trailing_json_list(value: Any) -> list[str]:
    text = str(value or "").strip()
    match = re.search(r",\s*(\[[^\]]*\])\s*$", text, re.S)
    if not match:
        return []
    return parse_json_list(match.group(1))


def parse_ret_sample(value: Any) -> tuple[str, list[str]]:
    text = str(value or "").strip()
    if not text:
        return "", []
    images = parse_trailing_json_list(text)
    answer = text
    match = re.search(r",\s*\[[^\]]*\]\s*$", text, re.S)
    if match:
        answer = text[: match.start()].strip()
    answer = unescape_unicode_literals(answer.strip().strip('"').strip())
    return answer, images


def find_summary_workbook() -> Path:
    matches = sorted(
        (path for path in DOWNLOADS.glob("*标准回答样例库*.xlsx") if not path.name.startswith("~$")),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not matches:
        raise FileNotFoundError("No standard answer example workbook found in Downloads.")
    return matches[0]


def find_review_workbook() -> Path | None:
    matches = sorted(
        (path for path in DOWNLOADS.glob("*逐题审查*.xlsx") if not path.name.startswith("~$")),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return matches[0] if matches else None


def find_teacher_reference_csv() -> Path | None:
    path = DOWNLOADS / "reference_teacher_v18_human_explicit_ret_on_v16.csv"
    if path.exists():
        return path
    matches = sorted(
        DOWNLOADS.glob("reference_teacher_v18*human*ret*.csv"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return matches[0] if matches else None


def load_summary_examples(path: Path) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    examples: dict[str, dict[str, Any]] = {}
    for row in read_sheet_rows(path, "标准样例库_350"):
        rid = compact(row.get("id"))
        if not rid:
            continue
        ret_answer, ret_images = parse_ret_sample(row.get("ret_sample"))
        image_ids = parse_json_list(row.get("image_ids_json")) or ret_images
        answer = compact(row.get("answer_sample_plain")) or compact(ret_answer)
        examples[rid] = {
            "id": rid,
            "question": compact(row.get("question")).strip('"'),
            "route_type": compact(row.get("route_type")),
            "manual_id": compact(row.get("manual_id")),
            "product": compact(row.get("product")),
            "language": compact(row.get("language")),
            "question_type": compact(row.get("question_type")),
            "answer_sample_plain": answer,
            "ret_sample": compact(row.get("ret_sample")),
            "image_ids": image_ids,
            "image_reason": compact(row.get("image_reason")),
            "forbidden_images": parse_json_list(row.get("forbidden_images")),
            "rule_note": compact(row.get("rule_note")),
            "source_basis": compact(row.get("source_basis")),
            "confidence": compact(row.get("confidence")),
            "status": compact(row.get("status")),
            "source": str(path),
            "audit_overrides": [],
        }

    policy_examples = []
    for row in read_sheet_rows(path, "非手册题"):
        rid = compact(row.get("id"))
        if not rid:
            continue
        policy_examples.append(
            {
                "id": rid,
                "question": compact(row.get("question")).strip('"'),
                "route_type": compact(row.get("route_type")),
                "product": compact(row.get("product")),
                "answer_sample_plain": compact(row.get("answer_sample_plain")),
                "note": compact(row.get("note")),
                "source": str(path),
            }
        )

    rule_patches = []
    for row in read_sheet_rows(path, "大模型规则补丁"):
        scope = compact(row.get("适用手册/范围"))
        trigger = compact(row.get("触发条件"))
        instruction = compact(row.get("规则补丁/给模型的指令"))
        if scope or trigger or instruction:
            rule_patches.append({"scope": scope, "trigger": trigger, "instruction": instruction, "source": str(path)})

    captions = []
    for row in read_sheet_rows(path, "图片caption索引"):
        image_id = compact(row.get("image_id"))
        if image_id:
            captions.append(
                {
                    "image_id": image_id,
                    "caption": compact(row.get("caption")),
                    "source": compact(row.get("source")) or str(path),
                    "used_count": compact(row.get("used_count")),
                }
            )
    return examples, policy_examples, rule_patches, captions


def normalize_header(header: str) -> str:
    return compact(header).lower().replace(" ", "").replace("_", "")


def pick(row: dict[str, Any], *keywords: str) -> Any:
    normalized = {normalize_header(key): key for key in row}
    for key_norm, key in normalized.items():
        if all(keyword.lower() in key_norm for keyword in keywords):
            return row.get(key)
    return None


def row_id(row: dict[str, Any]) -> str:
    for key in row:
        if normalize_header(key) in {"id", "题号"}:
            return compact(row.get(key))
    return compact(pick(row, "id"))


def extract_detail_rows(path: Path, trust_current_images: bool = False, include_current_answer: bool = False) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    details: list[dict[str, Any]] = []
    try:
        for ws in wb.worksheets:
            raw_rows = list(ws.iter_rows(values_only=True))
            for header_idx, raw_header in enumerate(raw_rows[:8]):
                header = [compact(value) for value in raw_header]
                norm = [normalize_header(value) for value in header]
                has_id = any(value in {"id", "题号"} for value in norm)
                has_question = any("题目" in value or "问题" in value or "question" in value for value in norm)
                if not has_id or not has_question:
                    continue
                for raw in raw_rows[header_idx + 1 :]:
                    if not any(value is not None and compact(value) for value in raw):
                        continue
                    row = {header[i]: raw[i] if i < len(raw) else None for i in range(len(header)) if header[i]}
                    rid = row_id(row)
                    if not rid:
                        continue
                    suggested_images = (
                        parse_json_list(pick(row, "建议", "图片"))
                        or parse_json_list(pick(row, "建议imageidsjson"))
                        or parse_json_list(pick(row, "建议", "image"))
                    )
                    current_images = parse_json_list(pick(row, "当前", "图片")) or parse_json_list(pick(row, "原图片"))
                    ret_value = pick(row, "建议ret") or pick(row, "ret_sample")
                    ret_answer, ret_images = parse_ret_sample(ret_value)
                    current_answer, current_answer_images = parse_ret_sample(
                        pick(row, "当前", "答案") or pick(row, "答案", "片段")
                    )
                    answer_hint = (
                        ret_answer
                        or (current_answer if include_current_answer else "")
                        or compact(pick(row, "建议", "答案"))
                        or compact(pick(row, "建议", "回答"))
                        or compact(pick(row, "建议", "要点"))
                        or compact(pick(row, "修改", "标准答案"))
                    )
                    if not suggested_images:
                        suggested_images = ret_images or current_answer_images
                    if not suggested_images and trust_current_images:
                        suggested_images = current_images
                    feedback = "；".join(
                        part
                        for part in (
                            compact(pick(row, "给大模型", "反馈")),
                            compact(pick(row, "问题说明")),
                            compact(pick(row, "正文")),
                            compact(pick(row, "文本")),
                            compact(pick(row, "内容核查")),
                            compact(pick(row, "文本正确性")),
                        )
                        if part
                    )
                    basis = "；".join(
                        part
                        for part in (
                            compact(pick(row, "依据")),
                            compact(pick(row, "依据章节")),
                            compact(pick(row, "核查依据")),
                            compact(pick(row, "修改理由")),
                            compact(pick(row, "图片caption预览")),
                        )
                        if part
                    )
                    details.append(
                        {
                            "id": rid,
                            "question": compact(pick(row, "题目") or pick(row, "问题") or pick(row, "question")).strip('"'),
                            "suggested_images": suggested_images,
                            "current_images": current_images,
                            "answer_hint": answer_hint,
                            "ret_sample": compact(ret_value),
                            "conclusion": compact(pick(row, "结论") or pick(row, "结论等级") or pick(row, "原答案结论") or pick(row, "当前答案结论")),
                            "text_review": feedback,
                            "image_review": compact(pick(row, "图片正确性") or pick(row, "图片判断") or pick(row, "图片结论")),
                            "action": compact(pick(row, "建议处理") or pick(row, "处理建议") or pick(row, "建议动作")),
                            "basis": basis,
                            "sheet": ws.title,
                            "source": str(path),
                        }
                    )
                break
    finally:
        wb.close()
    return details


def confidence_rank(text: str) -> int:
    order = {"高": 3, "中": 2, "低": 1}
    return order.get(text, 0)


def apply_detail_overrides(
    examples: dict[str, dict[str, Any]],
    detail_rows: list[dict[str, Any]],
    update_answers: bool = True,
    update_images: bool = True,
) -> dict[str, int]:
    by_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for detail in detail_rows:
        by_id[detail["id"]].append(detail)

    changed_answer = 0
    changed_images = 0
    for rid, details in by_id.items():
        if rid not in examples:
            continue
        example = examples[rid]
        # Prefer rows from explicit suggestion sheets, then rows that contain a concrete answer/image.
        def score(detail: dict[str, Any]) -> tuple[int, int, int]:
            sheet = detail.get("sheet", "")
            explicit_sheet = any(token in sheet for token in ("修正", "建议", "逐题", "校验", "核查"))
            return (
                1 if explicit_sheet else 0,
                1 if detail.get("suggested_images") is not None else 0,
                len(detail.get("answer_hint") or ""),
            )

        details = sorted(details, key=score, reverse=True)
        audit_notes = []
        for detail in details:
            note = {
                "source": detail["source"],
                "sheet": detail["sheet"],
                "conclusion": detail.get("conclusion", ""),
                "text_review": detail.get("text_review", ""),
                "image_review": detail.get("image_review", ""),
                "action": detail.get("action", ""),
                "basis": detail.get("basis", ""),
                "current_images": detail.get("current_images", []),
                "suggested_images": detail.get("suggested_images", []),
            }
            audit_notes.append(note)
            if update_images and detail.get("suggested_images") is not None and detail.get("suggested_images") != []:
                if example.get("image_ids") != detail["suggested_images"]:
                    example["image_ids"] = detail["suggested_images"]
                    changed_images += 1
                    example["image_reason"] = detail.get("basis") or detail.get("image_review") or example.get("image_reason", "")
                    example["source_basis"] = (example.get("source_basis", "") + "；人工样例核查报告覆盖图片").strip("；")
            elif update_images and detail.get("suggested_images") == [] and re.search(r"无需图片|不需要图片|不应配图|无图题|图片应为空|保持无图", " ".join(str(detail.get(k, "")) for k in ("image_review", "action", "basis", "text_review"))):
                if example.get("image_ids"):
                    example["image_ids"] = []
                    changed_images += 1
                    example["image_reason"] = detail.get("basis") or detail.get("image_review") or ""
            if update_answers and detail.get("answer_hint") and len(detail["answer_hint"]) >= 20:
                if compact(example.get("answer_sample_plain")) != detail["answer_hint"]:
                    example["answer_sample_plain"] = detail["answer_hint"]
                    changed_answer += 1
                    if detail.get("ret_sample"):
                        example["ret_sample"] = detail["ret_sample"]
        example["audit_overrides"] = audit_notes[:5]
        if audit_notes:
            example["source_basis"] = (example.get("source_basis", "") + "；已合并人工逐题核查报告").strip("；")
    return {"changed_answer": changed_answer, "changed_images": changed_images, "detail_ids": len(by_id)}


def load_teacher_reference_rets(path: Path | None) -> dict[str, dict[str, Any]]:
    if path is None or not path.exists():
        return {}
    rows: dict[str, dict[str, Any]] = {}
    with path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            rid = compact(row.get("id"))
            ret = row.get("ret") or ""
            answer, images = parse_ret_sample(ret)
            if rid and (answer or images):
                rows[rid] = {
                    "id": rid,
                    "answer": answer,
                    "image_ids": images,
                    "ret": ret,
                    "source": str(path),
                }
    return rows


def apply_teacher_reference_overrides(
    examples: dict[str, dict[str, Any]],
    teacher_rows: dict[str, dict[str, Any]],
) -> dict[str, int]:
    changed_answer = 0
    changed_images = 0
    for rid, ref in teacher_rows.items():
        example = examples.get(rid)
        if not example:
            continue
        answer = compact(ref.get("answer"))
        images = [str(image_id) for image_id in (ref.get("image_ids") or []) if str(image_id).strip()]
        if answer and compact(example.get("answer_sample_plain")) != answer:
            example["answer_sample_plain"] = answer
            changed_answer += 1
        if images and example.get("image_ids") != images:
            example["image_ids"] = images
            changed_images += 1
        if ref.get("ret"):
            example["ret_sample"] = compact(ref["ret"])
        example["teacher_reference_source"] = ref.get("source", "")
        example["source_basis"] = (example.get("source_basis", "") + "；已合并v18人工参考ret").strip("；")
    return {
        "teacher_reference_rows": len(teacher_rows),
        "teacher_changed_answer": changed_answer,
        "teacher_changed_images": changed_images,
    }


def load_rule_patches_from_workbook(path: Path | None) -> list[dict[str, Any]]:
    if path is None or not path.exists():
        return []
    try:
        rows = read_sheet_rows(path, "大模型规则补丁")
    except KeyError:
        return []
    patches = []
    for row in rows:
        scope = compact(row.get("适用手册/范围"))
        trigger = compact(row.get("触发条件"))
        instruction = compact(row.get("规则补丁/给模型的指令"))
        if scope or trigger or instruction:
            patches.append({"scope": scope, "trigger": trigger, "instruction": instruction, "source": str(path)})
    return patches


def annotate_example_quality(examples: dict[str, dict[str, Any]]) -> dict[str, int]:
    aligned = 0
    mismatched = 0
    image_without_pic = 0
    for example in examples.values():
        answer = str(example.get("answer_sample_plain") or "")
        image_ids = [str(image_id) for image_id in (example.get("image_ids") or []) if str(image_id).strip()]
        pic_count = answer.count("<PIC>")
        example["sample_pic_count"] = pic_count
        example["sample_image_count"] = len(image_ids)
        example["pic_image_aligned"] = pic_count == len(image_ids)
        if example["pic_image_aligned"]:
            aligned += 1
        else:
            mismatched += 1
            if image_ids and pic_count == 0:
                image_without_pic += 1
    return {
        "pic_image_aligned": aligned,
        "pic_image_mismatched": mismatched,
        "image_without_pic_examples": image_without_pic,
    }


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_sft(path: Path, examples: list[dict[str, Any]]) -> None:
    system = (
        "You are a multimodal customer-service manual agent. Answer in the user's language, "
        "use only manual evidence, preserve required warnings/notes/steps, and place <PIC> "
        "exactly where the selected images should appear."
    )
    rows = []
    for example in examples:
        if example.get("confidence") == "低":
            continue
        if example.get("pic_image_aligned") is False:
            continue
        answer = compact(example.get("answer_sample_plain"))
        if not answer:
            continue
        rows.append(
            {
                "messages": [
                    {"role": "system", "content": system},
                    {
                        "role": "user",
                        "content": (
                            f"Question: {example.get('question')}\n"
                            f"Manual: {example.get('manual_id')}\n"
                            f"Question type: {example.get('question_type')}\n"
                            f"Selected images: {example.get('image_ids')}\n"
                            f"Image rationale: {example.get('image_reason')}\n"
                            f"Rules: {example.get('rule_note')}"
                        ),
                    },
                    {"role": "assistant", "content": answer},
                ],
                "metadata": {
                    "id": example.get("id"),
                    "manual_id": example.get("manual_id"),
                    "language": example.get("language"),
                    "question_type": example.get("question_type"),
                    "image_ids": example.get("image_ids"),
                },
            }
        )
    write_jsonl(path, rows)


def main() -> None:
    summary_path = find_summary_workbook()
    examples_by_id, policy_examples, rule_patches, captions = load_summary_examples(summary_path)
    detail_paths = [path for path in sorted(REPORT_DIR.rglob("*.xlsx")) if not path.name.startswith("~$")]
    detail_rows: list[dict[str, Any]] = []
    for path in detail_paths:
        detail_rows.extend(extract_detail_rows(path))
    merge_stats = apply_detail_overrides(examples_by_id, detail_rows)

    teacher_path = find_teacher_reference_csv()
    teacher_rows = load_teacher_reference_rets(teacher_path)
    teacher_stats = apply_teacher_reference_overrides(examples_by_id, teacher_rows)

    review_path = find_review_workbook()
    review_rows = extract_detail_rows(review_path, trust_current_images=True, include_current_answer=False) if review_path else []
    review_stats = apply_detail_overrides(examples_by_id, review_rows, update_answers=False, update_images=True)
    detail_rows.extend(review_rows)
    rule_patches.extend(load_rule_patches_from_workbook(review_path))
    quality_stats = annotate_example_quality(examples_by_id)

    examples = sorted(examples_by_id.values(), key=lambda row: int(row["id"]) if str(row["id"]).isdigit() else str(row["id"]))
    write_jsonl(ASSET_DIR / "human_example_library.jsonl", examples)
    write_jsonl(ASSET_DIR / "human_policy_examples.jsonl", policy_examples)
    write_jsonl(ASSET_DIR / "human_rule_patches.jsonl", rule_patches)
    write_jsonl(ASSET_DIR / "human_caption_index.jsonl", captions)
    write_jsonl(ASSET_DIR / "human_audit_detail_rows.jsonl", detail_rows)
    write_sft(ASSET_DIR / "human_sft_messages.jsonl", examples)

    stats = {
        "summary_workbook": str(summary_path),
        "manual_examples": len(examples),
        "policy_examples": len(policy_examples),
        "rule_patches": len(rule_patches),
        "captions": len(captions),
        "detail_reports": len(detail_paths),
        "review_workbook": str(review_path) if review_path else "",
        "detail_rows": len(detail_rows),
        **merge_stats,
        **teacher_stats,
        "review_changed_images": review_stats["changed_images"],
        "review_detail_ids": review_stats["detail_ids"],
        **quality_stats,
        "by_confidence": Counter(example.get("confidence", "") for example in examples),
        "by_language": Counter(example.get("language", "") for example in examples),
        "by_manual": Counter(example.get("manual_id", "") for example in examples),
    }
    stats_path = ASSET_DIR / "human_example_library_stats.json"
    stats_path.write_text(json.dumps(stats, ensure_ascii=False, indent=2, default=dict), encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2, default=dict))


if __name__ == "__main__":
    main()
