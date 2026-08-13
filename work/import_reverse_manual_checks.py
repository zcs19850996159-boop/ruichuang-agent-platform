from __future__ import annotations

import json
import os
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
OUT_PATH = ROOT / "outputs" / "rag_assets" / "original_manual_reverse_checks.jsonl"


def rows_from_sheet(ws):
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        yield {
            headers[i]: row[i]
            for i in range(min(len(headers), len(row)))
            if headers[i]
        }


def main() -> None:
    xlsx_path = os.environ.get("REVERSE_CHECK_XLSX", "").strip()
    if not xlsx_path:
        raise SystemExit("Set REVERSE_CHECK_XLSX to the workbook path.")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    records = []

    for row in rows_from_sheet(wb["Summary"]):
        records.append(
            {
                "record_type": "manual_summary",
                "manual_id": str(row.get("Manual ID") or "").strip(),
                "product_cn": str(row.get("中文产品") or "").strip(),
                "source_candidate": str(row.get("原版/候选原版") or "").strip(),
                "confidence": str(row.get("匹配置信度") or "").strip(),
                "conclusion": str(row.get("反推结论") or "").strip(),
                "action": str(row.get("入库动作") or "").strip(),
                "source_url": str(row.get("source_url") or "").strip(),
            }
        )

    for row in rows_from_sheet(wb["Caption_Updates"]):
        image_id = str(row.get("Image ID / PIC") or "").strip()
        records.append(
            {
                "record_type": "caption_update",
                "manual_id": str(row.get("Manual ID") or "").strip(),
                "image_id": image_id,
                "caption_cn": str(row.get("反推后 caption_cn") or "").strip(),
                "evidence_source": str(row.get("证据来源") or "").strip(),
                "action": str(row.get("动作") or "").strip(),
                "risk_level": str(row.get("风险等级") or "").strip(),
                "notes": str(row.get("备注") or "").strip(),
            }
        )

    for row in rows_from_sheet(wb["Source_Evidence"]):
        records.append(
            {
                "record_type": "source_evidence",
                "manual_id": str(row.get("Manual ID") or "").strip(),
                "source_title": str(row.get("Source title") or "").strip(),
                "url": str(row.get("URL") or "").strip(),
                "key_evidence": str(row.get("关键匹配证据") or "").strip(),
                "supported_action": str(row.get("可支持的反推动作") or "").strip(),
            }
        )

    for row in rows_from_sheet(wb["Pending"]):
        records.append(
            {
                "record_type": "pending",
                "manual_id": str(row.get("Manual ID") or "").strip(),
                "product": str(row.get("产品") or "").strip(),
                "status": str(row.get("当前状态") or "").strip(),
                "next_search": str(row.get("下一步搜索策略") or "").strip(),
                "affects_ingestion": str(row.get("是否会影响当前入库") or "").strip(),
            }
        )

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUT_PATH.open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    print(f"wrote {len(records)} records to {OUT_PATH}")


if __name__ == "__main__":
    main()
